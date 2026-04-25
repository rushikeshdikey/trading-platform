"""Zerodha Console tradebook → Trade + Pyramid + Exit reconstruction.

Zerodha gives one row per execution (buy or sell leg). A "trade" in the journal
sense is the full round trip: an opening leg, optional same-side pyramids, and
one or more scale-out exits. We reconstruct by walking executions per symbol
chronologically and doing FIFO matching.

Supports both CSV and XLSX exports. Accepts older and newer column layouts by
normalising header names (case-insensitive, strip spaces/underscores).

Imports are append-only and deduplicated on Zerodha's per-execution ``Trade
ID``. New executions for a symbol that already has an open trade extend that
trade (pyramid for same-side, exit for opposite-side) rather than starting a
fresh trade — so manually-edited SL/setup/notes on existing trades survive.
"""
from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

import openpyxl
from sqlalchemy.orm import Session

from . import calculations as calc
from . import masterlist
from .models import Exit, ImportedExecution, Pyramid, Trade


COLUMN_ALIASES = {
    "symbol": {"symbol", "tradingsymbol", "scrip", "scripname", "script"},
    "trade_date": {"tradedate", "date", "tradedatetime", "orderexecutiontime"},
    "trade_type": {"tradetype", "buysell", "type", "transactiontype"},
    "quantity": {"quantity", "qty"},
    "price": {"price", "tradeprice", "averageprice", "rate"},
    # Zerodha tradebook has separate Trade ID (execution-unique) and Order ID
    # (shared across partial fills). Trade ID is the better dedup key; fall back
    # to Order ID only if Trade ID is absent.
    "trade_id": {"tradeid"},
    "order_id": {"orderid"},
    "exchange": {"exchange"},
}


@dataclass
class Execution:
    symbol: str
    date: date
    side: str  # 'B' or 'S'
    qty: int
    price: float
    trade_id: str | None = None
    order_id: str | None = None
    exchange: str | None = None


@dataclass
class ReconResult:
    trades_created: int = 0
    trades_extended: int = 0
    trades_closed: int = 0
    executions_parsed: int = 0
    executions_applied: int = 0
    executions_skipped_duplicate: int = 0
    skipped_rows: int = 0
    symbols_touched: set[str] = field(default_factory=set)
    warnings: list[str] = field(default_factory=list)


def _norm(key: str) -> str:
    return "".join(ch for ch in str(key).lower() if ch.isalnum())


def _resolve_columns(headers: list[str]) -> dict[str, str]:
    """Map our canonical names to the actual header in this file."""
    normalised = {_norm(h): h for h in headers if h is not None}
    resolved: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                resolved[canonical] = normalised[alias]
                break
    # Trade ID fallback → Order ID, so older exports without Trade ID still work.
    if "trade_id" not in resolved and "order_id" in resolved:
        resolved["trade_id"] = resolved["order_id"]
    return resolved


def _parse_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(s[: len(fmt) + 5], fmt).date() if len(s) > len(fmt) else datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_side(v) -> str | None:
    s = str(v or "").strip().lower()
    if s in ("buy", "b", "bought", "purchase"):
        return "B"
    if s in ("sell", "s", "sold"):
        return "S"
    return None


def _parse_float(v) -> float | None:
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(v) -> int | None:
    f = _parse_float(v)
    if f is None:
        return None
    return int(f)


def _row_to_execution(
    row: dict, cols: dict[str, str], row_idx: int, warnings: list[str]
) -> Execution | None:
    symbol = (str(row.get(cols["symbol"]) or "")).strip().upper()
    d = _parse_date(row.get(cols["trade_date"], ""))
    side = _parse_side(row.get(cols["trade_type"], ""))
    qty = _parse_int(row.get(cols["quantity"], ""))
    price = _parse_float(row.get(cols["price"], ""))

    if not (symbol and d and side and qty and price):
        warnings.append(
            f"row {row_idx}: skipped (symbol={symbol!r}, date={d}, side={side}, qty={qty}, price={price})"
        )
        return None

    tid = row.get(cols.get("trade_id", ""), "") if "trade_id" in cols else None
    oid = row.get(cols.get("order_id", ""), "") if "order_id" in cols else None
    ex = row.get(cols.get("exchange", ""), "") if "exchange" in cols else None
    return Execution(
        symbol=symbol,
        date=d,
        side=side,
        qty=qty,
        price=price,
        trade_id=str(tid).strip() if tid not in (None, "") else None,
        order_id=str(oid).strip() if oid not in (None, "") else None,
        exchange=str(ex).strip() if ex not in (None, "") else None,
    )


def parse_csv(content: str) -> tuple[list[Execution], list[str]]:
    """Return (executions, warnings). Tolerant of leading junk lines."""
    warnings: list[str] = []

    lines = content.splitlines()
    header_idx = 0
    for i, line in enumerate(lines[:25]):
        ll = line.lower()
        if ("symbol" in ll or "tradingsymbol" in ll) and ("quantity" in ll or "qty" in ll):
            header_idx = i
            break

    cleaned = "\n".join(lines[header_idx:])
    reader = csv.DictReader(StringIO(cleaned))
    if not reader.fieldnames:
        return [], ["CSV had no header row"]

    cols = _resolve_columns(list(reader.fieldnames))
    missing = [c for c in ("symbol", "trade_date", "trade_type", "quantity", "price") if c not in cols]
    if missing:
        return [], [f"CSV missing columns: {missing}. Headers present: {reader.fieldnames}"]

    execs: list[Execution] = []
    for row_idx, row in enumerate(reader, start=header_idx + 2):
        e = _row_to_execution(row, cols, row_idx, warnings)
        if e is not None:
            execs.append(e)
    return execs, warnings


def parse_xlsx(data: bytes | Path) -> tuple[list[Execution], list[str]]:
    """Parse a Zerodha Equity tradebook .xlsx.

    Handles the preamble rows Zerodha puts before the real header (client id,
    date range, blank rows). Picks the first sheet that contains a recognisable
    header row.
    """
    warnings: list[str] = []
    src = BytesIO(data) if isinstance(data, (bytes, bytearray)) else data
    try:
        wb = openpyxl.load_workbook(src, data_only=True)
    except Exception as e:
        return [], [f"Could not open xlsx: {e}"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(rows):
            lowered = [(_norm(v) if v is not None else "") for v in row]
            joined = " ".join(lowered)
            if "symbol" in joined and ("quantity" in joined or "qty" in joined):
                header_idx = i
                break
        if header_idx is None:
            continue

        header = [
            (str(v).strip() if v is not None else f"col{idx}")
            for idx, v in enumerate(rows[header_idx])
        ]
        cols = _resolve_columns(header)
        missing = [
            c
            for c in ("symbol", "trade_date", "trade_type", "quantity", "price")
            if c not in cols
        ]
        if missing:
            warnings.append(
                f"sheet {sheet_name!r}: missing columns {missing}. Saw {header}"
            )
            continue

        execs: list[Execution] = []
        for offset, raw in enumerate(rows[header_idx + 1 :], start=1):
            if all(v is None or str(v).strip() == "" for v in raw):
                continue
            row_dict = {header[idx]: raw[idx] for idx in range(min(len(header), len(raw)))}
            e = _row_to_execution(
                row_dict, cols, row_idx=header_idx + 1 + offset, warnings=warnings
            )
            if e is not None:
                execs.append(e)
        return execs, warnings

    return [], ["No sheet in workbook had a recognisable tradebook header"]


def parse_bytes(filename: str, raw: bytes) -> tuple[list[Execution], list[str]]:
    """Dispatch to csv / xlsx parser based on filename suffix."""
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return parse_xlsx(raw)
    if lower.endswith(".xls"):
        return [], [".xls (old Excel binary) isn't supported — re-download as .xlsx or .csv"]
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="replace")
    return parse_csv(text)


# -- Append + reconstruct --------------------------------------------------


def _open_trade_for(db: Session, symbol: str) -> Trade | None:
    return (
        db.query(Trade)
        .filter(Trade.instrument == symbol, Trade.status == "open")
        .order_by(Trade.entry_date.desc(), Trade.id.desc())
        .first()
    )


def _next_trade_no(db: Session) -> int:
    row = db.query(Trade.trade_no).order_by(Trade.trade_no.desc()).first()
    return int(row[0]) + 1 if row and row[0] else 1


def _new_trade(db: Session, symbol: str, leg: Execution, trade_no: int) -> Trade:
    t = Trade(
        trade_no=trade_no,
        instrument=symbol,
        side=leg.side,
        entry_date=leg.date,
        initial_entry_price=leg.price,
        initial_qty=leg.qty,
        sl=leg.price,  # Broker file has no SL — user sets it after import.
        status="open",
    )
    db.add(t)
    db.flush()
    return t


def apply_executions_append(
    db: Session, executions: list[Execution]
) -> ReconResult:
    """Apply each execution to the journal incrementally, skipping duplicates.

    Dedup key is the broker's Trade ID (falls back to Order ID when Trade ID
    isn't present). Executions whose id is already in `imported_executions` are
    skipped. New executions get applied in date+id order per symbol:

    - no open trade → open a new one
    - same-side leg → pyramid
    - opposite-side leg → exit (partial or full); if the close qty exceeds the
      open position, the trade closes and any overflow opens a new trade in
      the reverse direction.
    """
    result = ReconResult(executions_parsed=len(executions))

    # Build id-keyed map; de-dupe within the file (Zerodha sometimes exports
    # the same Trade ID twice if you pull overlapping date ranges).
    seen_in_file: set[str] = set()
    unique: list[Execution] = []
    without_id: list[Execution] = []
    for e in executions:
        if e.trade_id:
            if e.trade_id in seen_in_file:
                result.executions_skipped_duplicate += 1
                continue
            seen_in_file.add(e.trade_id)
            unique.append(e)
        else:
            without_id.append(e)  # will fall back to composite match

    if not unique and not without_id:
        return result

    # Skip anything already imported previously.
    existing_ids: set[str] = set()
    if unique:
        known = (
            db.query(ImportedExecution.trade_id)
            .filter(ImportedExecution.trade_id.in_([e.trade_id for e in unique]))
            .all()
        )
        existing_ids = {row[0] for row in known}

    to_apply = [e for e in unique if e.trade_id not in existing_ids]
    result.executions_skipped_duplicate += len(unique) - len(to_apply)

    # Same dedup shape for id-less legs: treat (symbol,date,side,qty,price) as key.
    def _composite_key(e: Execution) -> str:
        return f"noid|{e.symbol}|{e.date.isoformat()}|{e.side}|{e.qty}|{e.price}"

    if without_id:
        known_composites = {
            r[0]
            for r in db.query(ImportedExecution.trade_id)
            .filter(ImportedExecution.trade_id.like("noid|%"))
            .all()
        }
        for e in without_id:
            k = _composite_key(e)
            if k in known_composites:
                result.executions_skipped_duplicate += 1
            else:
                e.trade_id = k
                known_composites.add(k)
                to_apply.append(e)

    if not to_apply:
        return result

    # Group per symbol, apply chronologically.
    by_symbol: dict[str, list[Execution]] = defaultdict(list)
    for e in to_apply:
        by_symbol[e.symbol].append(e)

    next_no = _next_trade_no(db)

    for symbol, legs in by_symbol.items():
        legs.sort(key=lambda e: (e.date, e.trade_id or ""))
        current: Trade | None = _open_trade_for(db, symbol)
        position_qty: int = _signed_open_qty(current) if current else 0
        if current is not None:
            result.symbols_touched.add(symbol)

        for leg in legs:
            applied_to = _apply_leg(
                db, symbol, leg, current, position_qty, result, next_no
            )
            current, position_qty, next_no = applied_to

            db.add(
                ImportedExecution(
                    trade_id=leg.trade_id,  # guaranteed set above
                    symbol=symbol,
                    trade_date=leg.date,
                    side=leg.side,
                    qty=leg.qty,
                    price=leg.price,
                    order_id=leg.order_id,
                    exchange=leg.exchange,
                    applied_to_trade_id=current.id if current else None,
                )
            )
            result.executions_applied += 1
            result.symbols_touched.add(symbol)

    db.commit()
    return result


def _signed_open_qty(trade: Trade | None) -> int:
    if trade is None:
        return 0
    oq = calc.open_qty(trade)
    return oq if trade.side == "B" else -oq


def _apply_leg(
    db: Session,
    symbol: str,
    leg: Execution,
    current: Trade | None,
    position_qty: int,
    result: ReconResult,
    next_no: int,
) -> tuple[Trade, int, int]:
    """Apply one execution; return (current_trade, new_position_qty, next_trade_no)."""
    signed = leg.qty if leg.side == "B" else -leg.qty

    # No open trade → open one.
    if current is None or position_qty == 0:
        current = _new_trade(db, symbol, leg, next_no)
        result.trades_created += 1
        return current, signed, next_no + 1

    same_direction = (position_qty > 0 and leg.side == "B") or (
        position_qty < 0 and leg.side == "S"
    )
    if same_direction:
        seq = max((p.sequence for p in current.pyramids), default=0) + 1
        db.add(
            Pyramid(
                trade_id=current.id,
                sequence=seq,
                price=leg.price,
                qty=leg.qty,
                date=leg.date,
            )
        )
        result.trades_extended += 1
        return current, position_qty + signed, next_no

    # Opposite direction → exit (possibly closing + overflow).
    close_qty = min(abs(position_qty), leg.qty)
    seq = max((e.sequence for e in current.exits), default=0) + 1
    db.add(
        Exit(
            trade_id=current.id,
            sequence=seq,
            price=leg.price,
            qty=close_qty,
            date=leg.date,
        )
    )
    result.trades_extended += 1
    new_pos = position_qty + close_qty if position_qty < 0 else position_qty - close_qty
    if new_pos == 0:
        current.status = "closed"
        current.close_date = leg.date
        result.trades_closed += 1
        current = None

    overflow = leg.qty - close_qty
    if overflow > 0:
        overflow_leg = Execution(
            symbol=symbol,
            date=leg.date,
            side=leg.side,
            qty=overflow,
            price=leg.price,
            trade_id=leg.trade_id,
        )
        current = _new_trade(db, symbol, overflow_leg, next_no)
        result.trades_created += 1
        return current, overflow if leg.side == "B" else -overflow, next_no + 1

    return current, new_pos, next_no


# -- Public entry points ---------------------------------------------------


def fetch_today_via_kite(db: Session, user) -> ReconResult:
    """Pull today's executions from Kite's trades() API → same append pipeline.

    Kite's ``trades()`` returns every execution (fill) for the current trading
    day. We map each row to our ``Execution`` dataclass and hand it to
    ``apply_executions_append`` — same dedup ledger (Trade ID), same FIFO
    reconstruction as the xlsx upload path. Token-expiry is caller's problem:
    if not authed, raise so the UI can redirect to the Kite login flow.
    """
    from . import kite as kite_svc

    kc = kite_svc.client(user)
    if kc is None:
        raise RuntimeError("Not authenticated with Kite — log in via /account")

    try:
        rows = kc.trades() or []
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"Kite trades() failed: {exc}") from exc

    execs: list[Execution] = []
    for r in rows:
        side_raw = str(r.get("transaction_type") or "").upper()
        side = "B" if side_raw == "BUY" else "S" if side_raw == "SELL" else None
        if side is None:
            continue
        ts = r.get("fill_timestamp") or r.get("exchange_timestamp") or r.get("order_timestamp")
        d = ts.date() if hasattr(ts, "date") else _parse_date(ts)
        if d is None:
            continue
        try:
            qty = int(r.get("quantity") or 0)
            price = float(r.get("average_price") or 0.0)
        except (TypeError, ValueError):
            continue
        if qty <= 0 or price <= 0:
            continue
        tid = r.get("trade_id")
        execs.append(
            Execution(
                symbol=str(r.get("tradingsymbol") or "").upper(),
                date=d,
                side=side,
                qty=qty,
                price=price,
                trade_id=str(tid) if tid not in (None, "") else None,
                order_id=str(r.get("order_id") or "") or None,
                exchange=str(r.get("exchange") or "") or None,
            )
        )

    result = apply_executions_append(db, execs)
    result.executions_parsed = len(execs)
    return result


def import_tradebook(
    db: Session,
    *,
    csv_text: str | None = None,
    xlsx_bytes: bytes | None = None,
    filename: str | None = None,
    raw: bytes | None = None,
) -> ReconResult:
    """Append-import a Zerodha tradebook.

    Call with any one of:
      - ``csv_text`` (CSV string)
      - ``xlsx_bytes`` (raw xlsx)
      - ``filename`` + ``raw`` (dispatches by extension)
    """
    if filename is not None and raw is not None:
        executions, warnings = parse_bytes(filename, raw)
    elif xlsx_bytes is not None:
        executions, warnings = parse_xlsx(xlsx_bytes)
    elif csv_text is not None:
        executions, warnings = parse_csv(csv_text)
    else:
        raise ValueError("import_tradebook: no input provided")

    result = apply_executions_append(db, executions)
    result.warnings.extend(warnings)
    result.executions_parsed = len(executions)
    return result
