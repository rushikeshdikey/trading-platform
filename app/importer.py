"""Import existing Google-Sheets / Excel trading journal into the app's DB.

Parses the `DTrades` sheet from `Rushikesh's Trading Journal - 2025.xlsx`.
Safe to re-run: wipes existing trades before importing.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from . import masterlist, settings as app_settings
from .models import Exit, Pyramid, Trade


COLUMNS = {
    "trade_no": 1,
    "entry_date": 2,
    "instrument": 3,
    "strike": 4,
    "option_type": 5,
    "entry": 6,
    "avg_entry": 7,
    "sl": 8,
    "tsl": 9,
    "side": 10,
    "cmp": 11,
    "setup": 12,
    "base_duration": 13,
    "initial_qty": 14,
    "p1_price": 15,
    "p1_qty": 16,
    "p1_date": 17,
    "p2_price": 18,
    "p2_qty": 19,
    "p2_date": 20,
    "exit1_price": 24,
    "exit1_qty": 25,
    "exit1_date": 26,
    "exit2_price": 27,
    "exit2_qty": 28,
    "exit2_date": 29,
    "exit3_price": 30,
    "exit3_qty": 31,
    "exit3_date": 32,
    "position_status": 40,
    "close_date": 41,
    "plan_followed": 46,
    "exit_trigger": 47,
    "proficiency": 48,
    "growth_areas": 49,
    "observations": 50,
}


def _cell_value(ws, row, col):
    v = ws.cell(row, col).value
    return v


def _as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _as_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _as_int(v):
    f = _as_float(v)
    if f is None:
        return None
    return int(f)


def _as_str(v):
    if v is None:
        return None
    # openpyxl returns BSE numeric scrip codes (539195) as floats. Stringifying
    # yields "539195.0" — strip the trailing ".0" so downstream lookups work.
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def import_from_xlsx(db: Session, xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    if "DTrades" not in wb.sheetnames:
        return {"error": "DTrades sheet not found", "imported": 0}
    ws = wb["DTrades"]

    # Wipe existing trades (safer than trying to diff-merge).
    db.query(Exit).delete()
    db.query(Pyramid).delete()
    db.query(Trade).delete()
    db.commit()

    imported = 0
    skipped = 0
    for r in range(2, ws.max_row + 1):
        entry_date = _as_date(_cell_value(ws, r, COLUMNS["entry_date"]))
        instrument = _as_str(_cell_value(ws, r, COLUMNS["instrument"]))
        entry_price = _as_float(_cell_value(ws, r, COLUMNS["entry"]))
        qty = _as_int(_cell_value(ws, r, COLUMNS["initial_qty"]))
        side = _as_str(_cell_value(ws, r, COLUMNS["side"]))
        sl = _as_float(_cell_value(ws, r, COLUMNS["sl"]))
        if not (entry_date and instrument and entry_price and qty and side):
            skipped += 1
            continue

        # Indian sheets use 'B'/'S' for Buy/Sell; accept aliases too.
        side = side.strip().upper()[:1]
        if side not in ("B", "S"):
            skipped += 1
            continue

        if sl is None:
            sl = entry_price  # safety default so the NOT NULL column is happy

        pos_status = (_as_str(_cell_value(ws, r, COLUMNS["position_status"])) or "").lower()
        status = "closed" if pos_status == "closed" else "open"
        close_date = _as_date(_cell_value(ws, r, COLUMNS["close_date"]))

        trade = Trade(
            trade_no=_as_int(_cell_value(ws, r, COLUMNS["trade_no"])),
            instrument=instrument,
            strike=_as_float(_cell_value(ws, r, COLUMNS["strike"])),
            option_type=_as_str(_cell_value(ws, r, COLUMNS["option_type"])),
            side=side,
            entry_date=entry_date,
            initial_entry_price=entry_price,
            initial_qty=qty,
            sl=sl,
            tsl=_as_float(_cell_value(ws, r, COLUMNS["tsl"])),
            cmp=_as_float(_cell_value(ws, r, COLUMNS["cmp"])),
            setup=_as_str(_cell_value(ws, r, COLUMNS["setup"])),
            base_duration=_as_str(_cell_value(ws, r, COLUMNS["base_duration"])),
            status=status,
            close_date=close_date,
            plan_followed=_parse_bool(_cell_value(ws, r, COLUMNS["plan_followed"])),
            exit_trigger=_as_str(_cell_value(ws, r, COLUMNS["exit_trigger"])),
            proficiency=_as_str(_cell_value(ws, r, COLUMNS["proficiency"])),
            growth_areas=_as_str(_cell_value(ws, r, COLUMNS["growth_areas"])),
            observations=_as_str(_cell_value(ws, r, COLUMNS["observations"])),
        )

        for i, (pkey, qkey, dkey) in enumerate(
            [
                ("p1_price", "p1_qty", "p1_date"),
                ("p2_price", "p2_qty", "p2_date"),
            ],
            start=1,
        ):
            price = _as_float(_cell_value(ws, r, COLUMNS[pkey]))
            pqty = _as_int(_cell_value(ws, r, COLUMNS[qkey]))
            pdate = _as_date(_cell_value(ws, r, COLUMNS[dkey])) or entry_date
            if price and pqty:
                trade.pyramids.append(
                    Pyramid(sequence=i, price=price, qty=pqty, date=pdate)
                )

        for i, (pkey, qkey, dkey) in enumerate(
            [
                ("exit1_price", "exit1_qty", "exit1_date"),
                ("exit2_price", "exit2_qty", "exit2_date"),
                ("exit3_price", "exit3_qty", "exit3_date"),
            ],
            start=1,
        ):
            price = _as_float(_cell_value(ws, r, COLUMNS[pkey]))
            exq = _as_int(_cell_value(ws, r, COLUMNS[qkey]))
            exd = _as_date(_cell_value(ws, r, COLUMNS[dkey])) or close_date or entry_date
            if price and exq:
                trade.exits.append(
                    Exit(sequence=i, price=price, qty=exq, date=exd)
                )

        db.add(trade)
        imported += 1

    db.commit()

    # After import, make sure any setup/etc. values we encountered exist in MasterList
    _ensure_masterlist_covers_trades(db)

    return {"imported": imported, "skipped": skipped}


def _parse_bool(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "yes", "y", "1"):
        return True
    if s in ("false", "no", "n", "0"):
        return False
    return None


def _ensure_masterlist_covers_trades(db: Session) -> None:
    trades = db.query(Trade).all()
    seen = {c: set() for c in masterlist.CATEGORIES}
    for t in trades:
        if t.setup:
            seen["setup"].add(t.setup)
        if t.base_duration:
            seen["base_duration"].add(t.base_duration)
        if t.exit_trigger:
            seen["exit_trigger"].add(t.exit_trigger)
        if t.proficiency:
            for v in str(t.proficiency).split(","):
                v = v.strip()
                if v:
                    seen["proficiency"].add(v)
        if t.growth_areas:
            for v in str(t.growth_areas).split(","):
                v = v.strip()
                if v:
                    seen["growth_area"].add(v)
    for cat, vals in seen.items():
        for v in vals:
            masterlist.add_value(db, cat, v)


def import_capital_from_dashboard(db: Session, xlsx_path: Path) -> None:
    """Extract starting capital from the `DDashboard` Jul row as a baseline."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "DDashboard" not in wb.sheetnames:
        return
    ws = wb["DDashboard"]
    for r in range(2, min(ws.max_row, 20) + 1):
        starting = ws.cell(r, 3).value
        if isinstance(starting, (int, float)) and starting > 0:
            app_settings.set_value(db, "starting_capital", str(float(starting)))
            db.commit()
            return
